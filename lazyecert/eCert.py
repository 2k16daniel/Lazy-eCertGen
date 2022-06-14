from typing import Type
import click
import openpyxl

# Optional lists


@click.command()
@click.option('--config-file', type=click.File(), help='Path of your custom config file')
@click.option('--template', type=click.File(), help='Path of your eCertificate template ( Must be in image format)')
@click.option('--out-dir', default="output", type=click.Path(), help='Your desired output path')
@click.option('--excel-file', type=click.Path(), help='Path of your excel file')
@click.option('--coordinate-x', type=int, help='Coordinates on the certificate where will be printing the name ')
@click.option('--coordinate-y', type=int, help='Coordinates on the certificate where will be printing the name ')
def cli(out_dir, excel_file, config_file, template, coordinate_x, coordinate_y):
    """Stupid Cli app for bulk e-certificate generator and deployer via email."""
    # check first if config file exists
    if config_file is None:
        print('Config file is missing, skipping...')
    else:
        print(config_file)
    # load names from excel
    Name = excel_parser(excel_file)
    a, b = Name[3]
    print(b)


def excel_parser(excel_file):
    print('fetching names from your excel file...')
    obj = openpyxl.load_workbook(excel_file)
    sheet = obj.active
    Name = []
    # calculate num of rows that has value
    nrow_max = len([row for row in sheet if not all(
        [cell.value == None for cell in row])])
    #store all email and name into array (tupple)
    for i in range(2, nrow_max):
        Name.append([''.join(
            [
                sheet.cell(row=i, column=3).value,
                ' ',
                sheet.cell(row=i, column=4).value,
                ' ',
                sheet.cell(row=i, column=5).value
            ]
        ),
            sheet.cell(row=i, column=2).value])
    return Name
